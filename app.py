import os
import cv2
import openpyxl
import zipfile
from flask import Flask, render_template, request, send_file, after_this_request, jsonify, url_for
from tempfile import mkdtemp
from shutil import rmtree
from config import Config
from werkzeug.utils import secure_filename
import json
from PIL import Image, ImageDraw, ImageFont # type: ignore
from io import BytesIO
import pandas as pd
import pdf2image
from docx2pdf import convert
from pdf2docx import Converter
import traceback

app = Flask(__name__)
app.config.from_object(Config)

# Ensure required directories exist
for directory in ['static/uploads', 'temp', 'static/templates/SVIT-Exclusive', 'static/templates/General', 'static/fonts', 'static/css', 'static/js']:
    os.makedirs(directory, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in Config.ALLOWED_EXTENSIONS

def get_available_fonts():
    fonts = []
    fonts_dir = os.path.join(app.static_folder, 'fonts')
    for file in os.listdir(fonts_dir):
        if file.endswith(('.ttf', '.otf')):
            font_name = os.path.splitext(file)[0]
            font_web_path = f'fonts/{file}'  # Always use forward slash for web
            fonts.append({
                'name': font_name,
                'path': font_web_path
            })
    return fonts

@app.route('/')
def index():
    return render_template('welcome.html')

# General Routes
@app.route('/single-certificate-general')
def single_certificate_general():
    fonts = get_available_fonts()
    return render_template('single_certificate.html', section='general', fonts=fonts)

@app.route('/bulk-certificate-general')
def bulk_certificate_general():
    fonts = get_available_fonts()
    return render_template('bulk_certificate.html', section='general', fonts=fonts)

# SVIT Exclusive Routes
@app.route('/single-certificate-svit')
def single_certificate_svit():
    fonts = get_available_fonts()
    return render_template('single_certificate.html', section='svit', fonts=fonts)

@app.route('/bulk-certificate-svit')
def bulk_certificate_svit():
    fonts = get_available_fonts()
    return render_template('bulk_certificate.html', section='svit', fonts=fonts)

# Tools Routes
@app.route('/tools')
def tools():
    return render_template('tools.html')

@app.route('/tools/pdf-to-img')
def pdf_to_img():
    return render_template('tools/pdf_to_img.html')

@app.route('/tools/img-to-pdf')
def img_to_pdf():
    return render_template('tools/img_to_pdf.html')

@app.route('/tools/word-to-pdf')
def word_to_pdf():
    return render_template('tools/word_to_pdf.html')

@app.route('/tools/pdf-to-word')
def pdf_to_word():
    return render_template('tools/pdf_to_word.html')

@app.route('/tools/excel-to-pdf')
def excel_to_pdf():
    return render_template('tools/excel_to_pdf.html')

@app.route('/api/templates')
def get_templates():
    try:
        templates = []
        section = request.args.get('section', 'general')
        template_dir = os.path.join(app.static_folder, 'templates', 'SVIT-Exclusive' if section == 'svit' else 'General')
        
        print(f"Loading templates from: {template_dir}")
        
        # Create template directory if it doesn't exist
        os.makedirs(template_dir, exist_ok=True)
        
        # Get all files in the template directory
        template_files = [f for f in os.listdir(template_dir) if f.lower().endswith(('.jpg', '.png', '.jpeg'))]
        print(f"Found template files: {template_files}")
        
        # Process each template file
        for filename in template_files:
            # Get template ID and name from filename
            template_id = os.path.splitext(filename)[0]
            template_name = template_id.replace('-', ' ').title()
            
            # Add template to list
            templates.append({
                'id': template_id,
                'name': template_name,
                'path': url_for('static', filename=f'templates/{section}/{filename}')
            })
        
        print(f"Returning {len(templates)} templates")
        return jsonify(templates)
        
    except Exception as e:
        print(f"Error loading templates: {str(e)}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/template/<string:template_id>', methods=['GET'])
def get_template_preview(template_id):
    try:
        section = request.args.get('section', 'general')
        if section not in ['general', 'svit']:
            # Explicitly return JSON with 400 status
            return jsonify({'error': 'Invalid section specified'}), 400

        template_dir_name = 'SVIT-Exclusive' if section == 'svit' else 'General'
        template_dir = os.path.join(app.static_folder, 'templates', template_dir_name)

        # Check for both jpg and png extensions (case-insensitive)
        template_path = None
        for ext in ['.jpg', '.png', '.jpeg']:
            temp_path = os.path.join(template_dir, f'{template_id}{ext}')
            if os.path.exists(temp_path):
                template_path = temp_path
                break
            # Also check uppercase extensions just in case
            temp_path_upper = os.path.join(template_dir, f'{template_id}{ext.upper()}')
            if os.path.exists(temp_path_upper):
                 template_path = temp_path_upper
                 break

        if template_path:
            # Construct the URL for the static file
            filename = os.path.basename(template_path)
            image_url = url_for('static', filename=f'templates/{template_dir_name}/{filename}')
            print(f"Found template image: {filename}, URL: {image_url}")
            # Explicitly return JSON with 200 status
            return jsonify({'image_url': image_url}), 200
        else:
            print(f"Template image not found for ID: {template_id} in {template_dir}")
            # Explicitly return JSON with 404 status
            return jsonify({'error': f'Template image for ID "{template_id}" not found.'}), 404

    except Exception as e:
        print(f"Error fetching template preview: {str(e)}")
        traceback.print_exc()
        # Explicitly return JSON with 500 status
        return jsonify({'error': f'Failed to fetch template preview: {str(e)}'}), 500

@app.route('/api/generate-single', methods=['POST'])
def generate_single_certificate():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
            
        name = data.get('name')
        template_id = data.get('template_id')
        customizations = data.get('customizations', {})
        
        if not name or not template_id:
            return jsonify({'error': 'Name and template are required'}), 400
            
        # Get section from query parameters
        section = request.args.get('section', 'general')
        if section not in ['general', 'svit']:
            return jsonify({'error': 'Invalid section specified'}), 400
            
        # Check for both jpg and png extensions
        template_dir = os.path.join(app.static_folder, 'templates', 'SVIT-Exclusive' if section == 'svit' else 'General')
        template_path_jpg = os.path.join(template_dir, f'{template_id}.jpg')
        template_path_png = os.path.join(template_dir, f'{template_id}.png')
        
        template_path = None
        if os.path.exists(template_path_jpg):
            template_path = template_path_jpg
        elif os.path.exists(template_path_png):
            template_path = template_path_png
            
        if not template_path:
            return jsonify({'error': f'Template "{template_id}" not found in {template_dir}. Please ensure the template exists.'}), 404
            
        # Create certificate using PIL
        try:
            img = Image.open(template_path).convert('RGB')
        except Exception as e:
            # Print the specific error when opening image
            traceback.print_exc()
            return jsonify({'error': f'Failed to open template image: {str(e)}'}), 500
            
        draw = ImageDraw.Draw(img)
        width, height = img.size

        # Font size and color
        try:
            font_size = float(customizations.get('font_size', 1.5))
            base_font_size = min(width, height) * 0.08  # Base size is 8% of the smaller dimension
            pil_font_size = int(base_font_size * font_size)
            font_name = 'Arial'
            font_path = None
            font_error = None
            for ext in ['.ttf', '.otf']:
                temp_path = os.path.join(app.static_folder, 'fonts', f'{font_name}{ext}')
                if os.path.exists(temp_path):
                    font_path = temp_path
                    break
            if font_path:
                font = ImageFont.truetype(font_path, pil_font_size)
            else:
                font_error = f"Font '{font_name}' not found in static/fonts."
                raise FileNotFoundError(font_error)
        except Exception as e:
            # Print the specific error when loading font
            traceback.print_exc()
            font_error = font_error or str(e)
            return jsonify({'error': f'Font error: {font_error}'}), 500

        # Font color
        try:
            text_color = customizations.get('font_color', '#000000')
            text_color = tuple(int(text_color.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
        except Exception as e:
            # Print the specific error when processing color
            traceback.print_exc()
            return jsonify({'error': f'Invalid font color format: {str(e)}'}), 400

        # Position logic with improved scaling
        text_position = customizations.get('text_position', {})
        scale_x = width / 800  # Base scale for x
        scale_y = height / 600  # Base scale for y
        x_offset = int(float(text_position.get('x', 0)) * scale_x)
        y_offset = int(float(text_position.get('y', 0)) * scale_y)
        
        # Calculate text dimensions
        text_bbox = draw.textbbox((0, 0), name, font=font)
        text_width = text_bbox[2] - text_bbox[0]
        text_height = text_bbox[3] - text_bbox[1]
        
        # Center the text by default
        base_x = (width - text_width) // 2
        base_y = (height - text_height) // 2
        x = base_x + x_offset
        y = base_y + y_offset

        # Draw text with improved shadow
        shadow_color = (0, 0, 0)
        outline_thickness = max(2, pil_font_size // 20)
        for dx, dy in [(-2,-2), (-2,2), (2,-2), (2,2)]:
            draw.text((x+dx, y+dy), name, font=font, fill=shadow_color)
        draw.text((x, y), name, font=font, fill=text_color)

        # Save and return image
        img_io = BytesIO()
        img.save(img_io, 'PNG')
        img_io.seek(0)
        return send_file(
            img_io,
            mimetype='image/png',
            as_attachment=True,
            download_name=f'Certificate_{name.replace(" ", "_")}.png'
        )
    except Exception as e:
        # Print the full traceback for any other unexpected errors
        traceback.print_exc()
        print(f"Error generating certificate: {str(e)}")
        return jsonify({'error': f'Failed to generate certificate: {str(e)}'}), 500

@app.route('/api/generate-bulk', methods=['POST'])
def generate_bulk():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        template_id = request.form.get('template_id')
        customizations = json.loads(request.form.get('customizations', '{}'))
        
        if not file:
            return jsonify({'error': 'No file selected'}), 400
            
        if not file.filename:
            return jsonify({'error': 'No file selected'}), 400
            
        if not allowed_file(file.filename):
            return jsonify({'error': 'Invalid file type. Please upload an Excel file (.xlsx or .xls)'}), 400
            
        if not template_id:
            return jsonify({'error': 'Template ID is required'}), 400
        
        temp_dir = mkdtemp()
        try:
            # Save the uploaded file temporarily
            temp_excel_path = os.path.join(temp_dir, secure_filename(file.filename))
            file.save(temp_excel_path)
            
            # Process Excel file
            wb = openpyxl.load_workbook(temp_excel_path)
            sheet = wb.active
            generated_files = []
            
            # Check for both jpg and png extensions
            template_path_jpg = os.path.join(app.config['TEMPLATE_PATH'], f'{template_id}.jpg')
            template_path_png = os.path.join(app.config['TEMPLATE_PATH'], f'{template_id}.png')
            
            template_path = None
            if os.path.exists(template_path_jpg):
                template_path = template_path_jpg
            elif os.path.exists(template_path_png):
                template_path = template_path_png
                
            if not template_path:
                return jsonify({'error': f'Template {template_id} not found'}), 404
            
            # Check if Excel file has data
            if sheet.max_row < 2:
                return jsonify({'error': 'Excel file is empty or has no data'}), 400
            
            # Process each row
            for row in range(2, sheet.max_row + 1):
                name = sheet.cell(row=row, column=1).value
                if name:
                    try:
                        # Convert name to string and clean it
                        name = str(name).strip()
                        if name:  # Only process non-empty names
                            output_path = generate_certificate(name, template_path, temp_dir, customizations)
                            generated_files.append(output_path)
                    except Exception as e:
                        print(f"Error generating certificate for {name}: {str(e)}")
                        continue
            
            if not generated_files:
                return jsonify({'error': 'No certificates were generated. Please check your Excel file format.'}), 400
            
            # Create ZIP file
            zip_path = os.path.join(temp_dir, 'certificates.zip')
            with zipfile.ZipFile(zip_path, 'w') as zipf:
                for file_path in generated_files:
                    zipf.write(file_path, os.path.basename(file_path))
            
            @after_this_request
            def cleanup(response):
                try:
                    rmtree(temp_dir, ignore_errors=True)
                except Exception as e:
                    print(f"Error cleaning up temporary directory: {str(e)}")
                return response
            
            return send_file(
                zip_path,
                as_attachment=True,
                download_name='certificates.zip',
                mimetype='application/zip'
            )
            
        except Exception as e:
            print(f"Error processing bulk certificates: {str(e)}")
            rmtree(temp_dir, ignore_errors=True)
            return jsonify({'error': f'Error processing bulk certificates: {str(e)}'}), 500
            
    except Exception as e:
        print(f"Server error: {str(e)}")
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@app.route('/api/convert', methods=['POST'])
def convert_file():
    if 'file' not in request.files:
        return 'No file uploaded', 400
    
    file = request.files['file']
    conversion_type = request.form.get('type')
    
    if not file or not allowed_file(file.filename):
        return 'Invalid file', 400
    
    temp_dir = mkdtemp()
    try:
        # Handle different conversion types
        if conversion_type == 'pdf2img':
            return convert_pdf_to_image(file, temp_dir)
        elif conversion_type == 'img2pdf':
            return convert_image_to_pdf(file, temp_dir)
        elif conversion_type == 'word2pdf':
            return convert_word_to_pdf(file, temp_dir)
        elif conversion_type == 'pdf2word':
            return convert_pdf_to_word(file, temp_dir)
        elif conversion_type == 'excel2pdf':
            return convert_excel_to_pdf(file, temp_dir)
        else:
            return 'Invalid conversion type', 400
    except Exception as e:
        rmtree(temp_dir, ignore_errors=True)
        return str(e), 400

def add_logo(img, logo_path):
    # Implementation for adding logo
    pass

def add_signature(img, signature_path):
    # Implementation for adding signature
    pass

def convert_pdf_to_image(file, temp_dir):
    try:
        # Save the uploaded PDF file
        pdf_path = os.path.join(temp_dir, secure_filename(file.filename))
        file.save(pdf_path)
        
        # Convert PDF to images
        images = pdf2image.convert_from_path(pdf_path)
        
        # Create a ZIP file to store all images
        zip_path = os.path.join(temp_dir, 'converted_images.zip')
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for i, image in enumerate(images):
                # Save each page as PNG
                image_path = os.path.join(temp_dir, f'page_{i+1}.png')
                image.save(image_path, 'PNG')
                zipf.write(image_path, f'page_{i+1}.png')
        
        @after_this_request
        def cleanup(response):
            rmtree(temp_dir, ignore_errors=True)
            return response
            
        return send_file(
            zip_path,
            as_attachment=True,
            download_name='converted_images.zip',
            mimetype='application/zip'
        )
    except Exception as e:
        return str(e), 500

def convert_image_to_pdf(file, temp_dir):
    try:
        # Save the uploaded image
        image_path = os.path.join(temp_dir, secure_filename(file.filename))
        file.save(image_path)
        
        # Open the image and convert to PDF
        image = Image.open(image_path)
        pdf_path = os.path.join(temp_dir, 'converted.pdf')
        image.save(pdf_path, 'PDF')
        
        @after_this_request
        def cleanup(response):
            rmtree(temp_dir, ignore_errors=True)
            return response
            
        return send_file(
            pdf_path,
            as_attachment=True,
            download_name='converted.pdf',
            mimetype='application/pdf'
        )
    except Exception as e:
        return str(e), 500

def convert_word_to_pdf(file, temp_dir):
    try:
        # Save the uploaded Word file
        docx_path = os.path.join(temp_dir, secure_filename(file.filename))
        file.save(docx_path)
        
        # Convert to PDF
        pdf_path = os.path.join(temp_dir, 'converted.pdf')
        convert(docx_path, pdf_path)
        
        @after_this_request
        def cleanup(response):
            rmtree(temp_dir, ignore_errors=True)
            return response
            
        return send_file(
            pdf_path,
            as_attachment=True,
            download_name='converted.pdf',
            mimetype='application/pdf'
        )
    except Exception as e:
        return str(e), 500

def convert_pdf_to_word(file, temp_dir):
    try:
        # Save the uploaded PDF file
        pdf_path = os.path.join(temp_dir, secure_filename(file.filename))
        file.save(pdf_path)
        
        # Convert PDF to Word
        docx_path = os.path.join(temp_dir, 'converted.docx')
        cv = Converter(pdf_path)
        cv.convert(docx_path)
        cv.close()
        
        @after_this_request
        def cleanup(response):
            rmtree(temp_dir, ignore_errors=True)
            return response
            
        return send_file(
            docx_path,
            as_attachment=True,
            download_name='converted.docx',
            mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
    except Exception as e:
        return str(e), 500

def convert_excel_to_pdf(file, temp_dir):
    try:
        # Save the uploaded Excel file
        excel_path = os.path.join(temp_dir, secure_filename(file.filename))
        file.save(excel_path)
        
        # Read Excel file
        df = pd.read_excel(excel_path)
        
        # Create a new Excel file with better formatting
        writer = pd.ExcelWriter(os.path.join(temp_dir, 'formatted.xlsx'), engine='openpyxl')
        df.to_excel(writer, index=False)
        
        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        # Save the formatted Excel file
        writer.close()
        
        # Convert to PDF using docx2pdf
        pdf_path = os.path.join(temp_dir, 'converted.pdf')
        convert(os.path.join(temp_dir, 'formatted.xlsx'), pdf_path)
        
        @after_this_request
        def cleanup(response):
            rmtree(temp_dir, ignore_errors=True)
            return response
            
        return send_file(
            pdf_path,
            as_attachment=True,
            download_name='converted.pdf',
            mimetype='application/pdf'
        )
    except Exception as e:
        return str(e), 500

if __name__ == '__main__':
    app.run(debug=True)